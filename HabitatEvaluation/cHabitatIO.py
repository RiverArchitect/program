#!/usr/bin/python
try:
    import os, sys, logging, shutil
    # load routines from LifespanDesign
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\riverpy\\")
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\openpyxl\\")
    import fGlobal as fg
    import openpyxl as oxl  # modified package
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")


class Read:
    def __init__(self, full_file_name):
        # type defines lines to read in .inp file
        self.xlsx_file = full_file_name
        self.logger = logging.getLogger("habitat_evaluation")
        self.open_wb()

    def close_wb(self):
        try:
            self.wb.close()
        except:
            pass

    def col_name_to_num(self, letters):
        pow = 1
        col_int = 0
        for letter in letters[::-1]:
            col_int += (int(letter, 36) - 9) * pow
            pow *= 26
        return col_int + 64

    def col_num_to_name(self, col_int):
        col_int -= 64
        letters = ''
        while col_int:
            mod = (col_int - 1) % 26
            letters += chr(mod + 65)
            col_int = (col_int - 1) // 26
        return ''.join(reversed(letters))

    def open_wb(self):
        try:
            self.wb = oxl.load_workbook(filename=self.xlsx_file, read_only=True, data_only=True)
        except:
            self.wb = ""
            self.logger.info("ERROR: Failed to access " + str(self.xlsx_file).split("\\")[-1] + ".")
        try:
            self.ws = self.wb.worksheets[0]
        except:
            self.ws = []
            self.logger.info("ERROR: Could not find sheet.")

    def read_column(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        self.logger.info(
            "   * Reading data column from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(column) + str(
                start_row) + ") ...")

        data = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                valid_content = False
            if valid_content:
                data.append(cell_value)
                __row__ += 1
        return data

    def read_row_str(self, row, start_col, **kwargs):
        # reads ROW beginning at COL until it meets an empty cell
        # row = INT
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
                if "end_col" in opt_var[0]:
                    # last relevant column
                    end_col = opt_var[1]
                if "if_row" in opt_var[0]:
                    # row that tells if the content of the actual cell is relevant
                    if_row = opt_var[1]
        except:
            pass

        if not ("col_skip" in locals()):
            col_skip = 1
        if not ("end_col" in locals()):
            end_col = "XFA"  # assumption: column XFA is the last that can be handled within a spreadsheet
        if not ("if_row" in locals()):
            if_row = row

        self.logger.info(
            "   * Reading string row from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(start_col) + str(row) + ") ...")

        str_data = []
        valid_content = True
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        while valid_content:
            cell_string = str(self.ws[str(__col__) + str(row)].value)
            valid_content = self.test_cell_content(str(__col__) + str(if_row))
            if valid_content:
                str_data.append(cell_string)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
            else:
                valid_content = False
        return str_data

    def test_cell_content(self, cell_str):
        # cell_str = str, e.g., "C4"
        if str(self.ws[cell_str].value) == "None":
            return False
        else:
            return True

    def __call__(self):
        print("Class Info: <type> = Read (Module: HabitatEnhancement)")


class Write:
    def __init__(self):
        # type(condition) ==  STR of folder name containing HSI rasters
        self.chsi_ras = []
        self.condition = ""
        self.dict_Q_h_ras = {}
        self.dict_Q_u_ras = {}
        self.dir_in_ras = r""
        self.dir_xlsx_out = os.path.dirname(os.path.abspath(__file__)) + "\\AUA\\"

        fg.chk_dir(self.dir_xlsx_out)

        self.logger = logging.getLogger("habitat_evaluation")
        self.discharges = []
        self.h_rasters = []
        self.u_rasters = []

        # dummy workbook variable instantiations
        self.xlsx_template = os.path.dirname(os.path.abspath(__file__)) + "\\.templates\\empty.xlsx"
        self.wb = oxl.load_workbook(filename=self.xlsx_template)
        self.wb_out_name = ""
        self.ws = self.wb.worksheets[0]

        # define columns
        self.col_Q = "B"
        self.col_ras_h = "C"
        self.col_ras_u = "D"

        # define rows
        self.row_start = 4

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def copy_wb(self, full_wb_inpath, full_save_path):
        try:
            self.logger.info("   * Creating a workbook copy: ")
            wb = oxl.load_workbook(filename=full_wb_inpath)
            self.logger.info("     " + full_save_path)
            wb.save(full_save_path)
            wb.close()
            self.logger.info("   * OK")
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def make_condition_xlsx(self, fish_sn):
        # type(cover_applies) ==  BOOL
        # fish_sn == STR -- 3 digits indicating fish species and lifestage

        ras_name_list = [rn for rn in os.listdir(self.dir_in_ras) if os.path.isdir(os.path.join(self.dir_in_ras, rn))]
        if ras_name_list.__len__() < 1:
            # look for geoTIFFs if ras_name_list is empty
            ras_name_list = [i for i in os.listdir(self.dir_in_ras) if i.endswith('.tif')]

        self.logger.info("   * Analyzing relevant discharges and matching rasters ...")
        for rn in ras_name_list:
            if rn[0] == "h":
                self.logger.info("     -- Found flow depth raster: " + str(rn))
                try:
                    if str(rn).endswith("k") or str(rn).endswith("k.tif"):
                        # multiply "k"-raster name discharges with 1000
                        thousand = 1000.0
                    else:
                        thousand = 1.0
                    _Q_ = float(str(rn).split("h")[1].split(".tif")[0].split("k")[0]) * thousand
                except:
                    _Q_ = 0.0
                self.discharges.append(_Q_)
                self.h_rasters.append(rn)
            if rn[0] == "u":
                self.logger.info("     -- Found flow velocity raster: " + str(rn))
                self.u_rasters.append(rn)

        # make Q_flowdur-raster dictionary
        self.dict_Q_h_ras = dict(zip(self.discharges, self.h_rasters))
        self.dict_Q_u_ras = dict(zip(self.discharges, self.u_rasters))
        # order discharges in descending order for flow duration curve
        self.discharges.sort(reverse=True)

        # open relevant workbook
        self.wb_out_name = self.dir_xlsx_out + str(self.condition) + "_" + fish_sn + ".xlsx"
        if os.path.isfile(self.wb_out_name):
            try:
                self.logger.info("   * creating safety copy of existing workbook ...")
                shutil.copy2(self.wb_out_name, self.dir_xlsx_out + str(self.condition) + "_" + fish_sn + "_old.xlsx")
                fg.rm_file(self.wb_out_name)
            except:
                self.logger.info("ERROR: Failed to access " + str(self.wb_out_name) + ".")
                return -1
        # else:
        try:
            self.logger.info("   * Creating new workbook (copy from .templates folder) ...")
            __wb__ = oxl.load_workbook(filename=self.xlsx_template)
        except:
            self.logger.info("ERROR: Failed to load " + str(self.xlsx_template) + ".")
            return -1
        self.wb = __wb__
        __wb__.close()
        try:
            self.ws = self.wb[str(self.wb.sheetnames[0])]
        except:
            self.logger.info("ERROR: TEMPLATE sheet does not exist.")
            return -1

        # write discharges and associated raster names
        self.logger.info("   * Writing discharges and associated raster names to workbook.")
        for iq in range(0, self.discharges.__len__()):
            try:
                self.ws[self.col_Q + str(self.row_start + iq)].value = self.discharges[iq]
                self.ws[self.col_ras_h + str(self.row_start + iq)].value = self.dict_Q_h_ras[self.discharges[iq]]
                self.ws[self.col_ras_u + str(self.row_start + iq)].value = self.dict_Q_u_ras[self.discharges[iq]]
            except:
                self.logger.info("ERROR: Invalid cell assignment for discharge / rasters.")
        self.logger.info("   * OK")
        return self.wb_out_name

    def open_wb(self, full_wb_path, ws):
        # full_wb_path = STR -- e.g., C:\\...\\my.xlsx
        # ws = STR or INT -- worksheet name or number in workbook
        self.logger.info("   * using " + str(full_wb_path))
        self.wb = oxl.load_workbook(filename=full_wb_path)
        try:
            ws = int(ws)
        except:
            pass
        if type(ws) == int:
            try:
                self.ws = self.wb[str(self.wb.sheetnames[ws])]
            except:
                pass
        else:
            self.ws = self.wb[ws]

    def read_column(self, column, start_read_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        # returns dictionary with cell data and cell information

        data = {}
        valid_content = True
        __row__ = start_read_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                valid_content = False
            if valid_content:
                data.update({cell_value: str(column) + str(__row__)})
                __row__ += 1
        self.logger.info("   * OK")
        return data

    def save_close_wb(self, *full_save_path):
        try:
            self.wb_out_name = full_save_path[0]
        except:
            pass

        try:
            self.logger.info("   * Saving as: " + self.wb_out_name)
            self.wb.save(self.wb_out_name)
            self.wb.close()
            self.logger.info("   * OK")
        except:
            self.logger.info("ERROR: Invalid file name or data.")


    def set_directories(self, condition, xlsx_template):
        self.condition = condition
        self.dir_in_ras = os.path.abspath(
            os.path.join(os.path.dirname(__file__), '..')) + "\\01_Conditions\\" + str(condition) + "\\"
        fg.chk_dir(self.dir_in_ras)

        self.xlsx_template = xlsx_template  # full file path
        self.logger.info("   * using workbook template: " + str(xlsx_template).split("\\")[-1])

    def write_data_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * ERROR: Could not write value to CELL " + str(column) + str(row))

    def write_data_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list must be a LIST object
        self.logger.info("   * Writing column data (starting at " + str(column) + str(start_row) + ") ...")
        __row__ = start_row
        for val in data_list:
            self.ws[str(column) + str(__row__)].value = val
            __row__ += 1
        self.logger.info("   * OK")

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (Module: cHabitatIO)")

