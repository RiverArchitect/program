# !/usr/bin/python
try:
    import sys, os, logging
    # make openypyxl accessible
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\openpyxl\\")
    import openpyxl as oxl
except:
    print("ExceptionERROR: Cannot find package files (/.site_packages/openpyxl/openpyxl).")


class ThresholdDirector:
    # Reads threshold values from file as a function of feature name
    def __init__(self, *args):
        self.logger = logging.getLogger("logfile")
        try:
            # LF instantiates with args[0] = True
            # avoids circular relations
            self.feat_id = args[0]
            sys.path.append(
                os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\riverpy\\")
            import cDefinitions as cdef  # contains reach and feature definitions
            self.features = cdef.Features()
        except:
            self.logger.info("ERROR: Invalid feature ID.")
            self.feat_id = ""
            self.features = ""

        self.thresh_row_dict = {"d2w_low": 7, "d2w_up": 8, "det_low": 9, "det_up": 10, "u": 12, "h": 11, "Fr": 13,
                                "D": 14, "freq": 15, "mu_bad": 16, "mu_good": 17, "mu_method": 18, "sf": 19,
                                "inverse_tcd": 21, "fill": 22, "scour": 23, "S0": 20, "taux": 6}
        self.unit_conv_candidates = ["d2w_low", "d2w_up", "det_low", "det_up", "u", "h", "D", "fill", "scour"]
        self.thresh_xlsx = os.path.dirname(os.path.abspath(__file__)) + "\\.templates\\threshold_values.xlsx"

        try:
            self.wb = oxl.load_workbook(filename=self.thresh_xlsx, read_only=True, data_only=True)
            wb_open = True
        except:
            wb_open = False
            self.logger.info("ERROR: Could not open threshold_values.xlsx.")
        try:
            if wb_open:
                self.ws = self.wb['thresholds']
            else:
                self.ws = ""
        except:
            self.logger.info("ERROR: Could not find sheet \'thresholds\' in threshold_values.xlsx.")

        # identify unit system
        try:
            unit_cell = self.ws["E28"].value
            if str(unit_cell).lower() == "u.s. customary":
                self.ft2m = 0.3047992
            else:
                self.ft2m = 1.0
        except:
            self.ft2m = 0.3047992
            self.logger.info("WARNING: Could not identify unit system settings (use default = U.S. customary).")

    def close_wb(self):
        self.wb.close()

    def get_thresh_value(self, thresh_type):
        try:
            cell_value = self.ws[str(self.features.thresh_col_dict[self.feat_id]) +
                                 str(self.thresh_row_dict[thresh_type])].value
        except:
            cell_value = -1
            self.logger.info("ERROR: Failed to read values from threshold_values.xlsx (return -1).")

        # convert value type
        info = cell_value
        try:
            if not(type(cell_value) == bool):
                cell_value = float(cell_value)
                if thresh_type in self.unit_conv_candidates:
                    cell_value = cell_value * self.ft2m
        except:
            if (str(cell_value).lower() == "true") or (str(cell_value).lower() == "false"):
                cell_value = bool(cell_value)
            else:
                try:
                    cell_value = str(cell_value)
                    if not(cell_value == "na"):
                        value_list = []
                        cell_value = str(cell_value).split(",")
                        for e in cell_value:
                           value_list.append(e.strip())
                        cell_value = value_list
                except:
                    self.logger.info("ERROR: Could not identify cell value type (tried: float, bool, str, list).")
        try:
            self.logger.info(
                "      Threshold/identifier for " + str(thresh_type) + ": " + str(info) + " " + str(type(cell_value)))
        except:
            self.logger.info("ERROR: No valid threshold for " + str(thresh_type) + " found. Check cell " + str(
                self.features.thresh_col_dict[self.feat_id]) + str(
                self.thresh_row_dict[thresh_type]) + " in threshold_values.xlsx.")
        return cell_value
