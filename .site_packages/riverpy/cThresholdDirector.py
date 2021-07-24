# !/usr/bin/python
try:
    import sys, os, logging
    import config
    sys.path.append(config.dir2oxl)
    import openpyxl as oxl
except:
    print("ExceptionERROR: Cannot find package files (.site_packages/openpyxl/openpyxl).")


class ThresholdDirector:
    # Reads threshold values from file as a function of feature name
    def __init__(self, *args):
        self.logger = logging.getLogger("logfile")
        try:
            # LF instantiates with args[0] = True
            # avoids circular relation and only locally import cDef
            import cDefinitions as cDef  # contains reach and feature definitions
            self.feat_id = args[0]
            self.features = cDef.FeatureDefinitions()
            self.feature_reader = cDef.FeatureReader()
        except:
            self.logger.info("ERROR: Invalid feature ID.")
            self.feat_id = ""
            self.features = ""
        self.all_row = {}
        self.thresh_row_dict = self.feature_reader.get_rows()
        # self.thresh_row_dict = {"d2w_low": 8, "d2w_up": 9, "det_low": 10, "det_up": 11, "u": 13, "h": 12, "Fr": 14,
        #                         "D": 15, "freq": 16, "mu_bad": 17, "mu_good": 18, "mu_method": 19, "sf": 20,
        #                         "inverse_tcd": 22, "fill": 23, "scour": 24, "S0": 21, "taux": 7, "lf": 25, "ds": 26,
        #                         "name": 4}
        self.unit_conv_candidates = ["d2w_low", "d2w_up", "det_low", "det_up", "u", "h", "D", "fill", "scour"]
        self.thresh_xlsx = config.xlsx_thresholds

        try:
            self.wb = oxl.load_workbook(filename=self.thresh_xlsx, read_only=True, data_only=True)
            wb_open = True
        except:
            wb_open = False
            self.logger.info("ERROR: Could not open threshold_values.xlsx.")
        try:
            if wb_open:
                self.ws = self.wb['thresholds']
            else:
                self.ws = ""
        except:
            self.logger.info("ERROR: Could not find sheet \'thresholds\' in threshold_values.xlsx.")

        # identify unit system
        try:
            unit_cell = self.ws.cell(row=self.thresh_row_dict['unit'], column=5).value
            if str(unit_cell).lower() == "u.s. customary":
                self.ft2m = 0.3047992
            else:
                self.ft2m = 1.0
        except:
            self.ft2m = 0.3047992
            self.logger.info("WARNING: Could not identify unit system settings (use default = U.S. customary).")

    def close_wb(self):
        self.wb.close()

    def get_thresh_value(self, thresh_type):
        try:
            cell_value = self.ws.cell(row=self.thresh_row_dict[thresh_type], column=self.features.thresh_col_dict[self.feat_id]).value
        except:
            cell_value = -1
            self.logger.info("ERROR: Failed to read values from threshold_values.xlsx (return -1).")

        # convert value type
        info = cell_value
        try:
            if not(type(cell_value) == bool):
                cell_value = float(cell_value)
                if thresh_type in self.unit_conv_candidates:
                    cell_value = cell_value * self.ft2m
        except:
            if (str(cell_value).lower() == "true") or (str(cell_value).lower() == "false"):
                cell_value = bool(cell_value)
            else:
                try:
                    cell_value = str(cell_value)
                    if not(cell_value == "na"):
                        value_list = []
                        cell_value = str(cell_value).split(",")
                        for e in cell_value:
                           value_list.append(e.strip())
                        cell_value = value_list
                except:
                    self.logger.info("ERROR: Could not identify cell value type (tried: float, bool, str, list).")
        try:
            self.logger.info(
                "      Threshold/identifier for " + str(thresh_type) + ": " + str(info) + " " + str(type(cell_value)))
        except:
            self.logger.info("ERROR: No valid threshold for " + str(thresh_type) + " found. Check cell " + str(
                self.features.thresh_col_dict[self.feat_id]) + str(
                self.thresh_row_dict[thresh_type]) + " in threshold_values.xlsx.")
        return cell_value
