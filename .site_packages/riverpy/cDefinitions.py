# !/usr/bin/python
import sys, os
try:
    sys.path.append(os.path.dirname(__file__))
    import config
    from collections import defaultdict
    import cReachManager as cRM
    sys.path.append(config.dir2oxl)
    import openpyxl as oxl
except:
    print("ExceptionERROR: Cannot find riverpy (%s)." % os.path.dirname(__file__))




class FeatureReader:
    # Reads threshold values from file as a function of feature name
    def __init__(self):
        self.main_dict = defaultdict(list)
        self.all_row = {}
        self.row_dict = {}
        self.row_feat_names = 4
        self.row_feat_ids = 5
        self.path2lf = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\..')) + "\\LifespanDesign\\"
        self.thresh_xlsx = self.path2lf + ".templates\\threshold_values.xlsx"
        try:
            self.wb = oxl.load_workbook(filename=self.thresh_xlsx, read_only=True, data_only=True)
            wb_open = True
        except:
            wb_open = False
            print("FeatureReader: Could not find " + str(self.thresh_xlsx))
        try:
            if wb_open:
                self.ws = self.wb['thresholds']
            else:
                self.ws = ""
        except:
            print("FeatureReader: Could not find sheet \'thresholds\' in threshold_values.xlsx.")

    def close_wb(self):
        self.wb.close()

    def get_rows(self):
        # Returns rows dynamically from threshold_values file
        max_row = self.ws.max_row
        for i in range(6, max_row + 1):
            cell_obj = self.ws.cell(row=i, column=2).value
            if not (cell_obj in self.all_row):
                self.all_row[cell_obj] = i
        self.row_dict['d2w_low'] = self.all_row['Depth to water table (min)']
        self.row_dict['d2w_up'] = self.all_row['Depth to water table (max)']
        self.row_dict['det_low'] = self.all_row['Detrended DEM (min)']
        self.row_dict['det_up'] = self.all_row['Detrended DEM (max)']
        self.row_dict['u'] = self.all_row['Flow velocity']
        self.row_dict['h'] = self.all_row['Flow depth']
        self.row_dict['Fr'] = self.all_row['Froude number']
        self.row_dict['D'] = self.all_row['Grain size']
        self.row_dict['freq'] = self.all_row['Design map frequency threshold']
        self.row_dict['mu_bad'] = self.all_row['Morphological Units: avoidance']
        self.row_dict['mu_good'] = self.all_row['Morphological Units: relevance']
        self.row_dict['mu_method'] = self.all_row['Morphological Units: application \n(0 = avoidance, 1 = relevance)']
        self.row_dict['sf'] = self.all_row['Safety factor']
        self.row_dict['inverse_tcd'] = self.all_row['Topographic change: inverse relevance']
        self.row_dict['scour'] = self.all_row['Topographic change: scour rate']
        self.row_dict['fill'] = self.all_row['Topographic change: fill rate']
        self.row_dict['S0'] = self.all_row['Terrain slope']
        self.row_dict['taux'] = self.all_row['Critical dimensionless bed shear stress']
        self.row_dict['lf'] = self.all_row['Apply Lifespan Mapping']
        self.row_dict['ds'] = self.all_row['Apply Design Mapping']
        self.row_dict['unit'] = self.all_row['CHOOSE UNIT SYSTEM:']
        self.row_dict['name'] = 4
        return self.row_dict

    def get_columns(self, name):
        # Returns columns dynamically for merged parent rows
        max_col = self.ws.max_column
        for i in range(5, max_col + 1):
            cell_obj = self.ws.cell(row=2, column=i).value
            if cell_obj != None:
                value = cell_obj
                if not (i in self.main_dict[value]):
                    self.main_dict[value].append(i)
            else:
                if not (i in self.main_dict[value]):
                    self.main_dict[value].append(i)
        return self.main_dict[name]

    def get_feat_id(self, column_list):
        feature_id_list = []
        for col in column_list:
            try:
                cell_value = str(self.ws.cell(row=self.row_feat_ids, column=col).value)
            except:
                cell_value = ""
                print("FeatureReader feature id: Failed to read values from threshold_values.xlsx (return empty).")
            feature_id_list.append(cell_value)
        return feature_id_list

    def get_feat_name(self, column_list):
        feature_list = []
        for col in column_list:
            try:
                cell_value = str(self.ws.cell(row=self.row_feat_names, column=col).value)
            except:
                cell_value = ""
                print("FeatureReader feature name: Failed to read values from threshold_values.xlsx (return empty).")
            feature_list.append(cell_value)
        return feature_list

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = FeatureReader (%s)" % os.path.dirname(__file__))
        print(dir(self))


class FeatureDefinitions:
    def __init__(self, *args):
        # args[0]: BOOL - False if called from application that cannot use "Custom DEM" option
        try:
            use_cust = args[0]
        except:
            use_cust = True  # Default value
        self.read_user_input = FeatureReader()
        # Modify feature/plantings IDs, names and columns with thresholds types in the next lines
        self.threshold_cols_framework = self.read_user_input.get_columns('TERRAFORMING')
        self.threshold_cols_plants = self.read_user_input.get_columns('VEGETATION SEEDLINGS/SAPLINGS')
        self.threshold_cols_toolbox = self.read_user_input.get_columns('NATURE-BASED (OTHER)')
        self.threshold_cols_complement = self.read_user_input.get_columns('CONNECTIVITY')
        self.fill_cols = [5, 8, 9, 25, 21, 23, 24]  # columns with terrain filling feature IDs
        self.excavate_cols = [5, 6, 7, 8, 9]  # columns with terrain lowering feature IDs

        # DO NOT MODIFY ANYTHING DOWN HERE
        self.id_list_framework = self.read_user_input.get_feat_id(self.threshold_cols_framework)
        self.id_list_plants = self.read_user_input.get_feat_id(self.threshold_cols_plants)
        self.id_list_toolbox = self.read_user_input.get_feat_id(self.threshold_cols_toolbox)
        self.id_list_complement = self.read_user_input.get_feat_id(self.threshold_cols_complement)

        self.name_list_framework = self.read_user_input.get_feat_name(self.threshold_cols_framework)
        self.name_list_plants = self.read_user_input.get_feat_name(self.threshold_cols_plants)
        self.name_list_toolbox = self.read_user_input.get_feat_name(self.threshold_cols_toolbox)
        self.name_list_complement = self.read_user_input.get_feat_name(self.threshold_cols_complement)

        self.fill_ids = self.read_user_input.get_feat_id(self.fill_cols)
        self.excavate_ids = self.read_user_input.get_feat_id(self.excavate_cols)
        self.read_user_input.close_wb()

        # merge lists
        self.id_list = self.id_list_framework + self.id_list_plants + self.id_list_toolbox + self.id_list_complement
        self.name_list = self.name_list_framework + self.name_list_plants + self.name_list_toolbox + self.name_list_complement
        self.threshold_cols = self.threshold_cols_framework + self.threshold_cols_plants + self.threshold_cols_toolbox + self.threshold_cols_complement

        self.feat_class_id_dict = {"terraforming": self.id_list_framework, "plantings": self.id_list_plants,
                                   "bioengineering": self.id_list_toolbox, "connectivity": self.id_list_complement}
        self.feat_class_name_dict = {"terraforming": self.name_list_framework, "plantings": self.name_list_plants,
                                     "bioengineering": self.name_list_toolbox, "connectivity": self.name_list_complement}

        if use_cust:
            self.id_list.append("cust")
            self.name_list.append("Custom DEM")
            self.threshold_cols.append("T")

        # create dictionaries
        self.feat_name_dict = dict(zip(self.id_list, self.name_list))
        self.feat_num_dict = dict(zip(self.id_list, range(0, self.id_list.__len__())))
        self.thresh_col_dict = dict(zip(self.id_list, self.threshold_cols))
        self.col_name_dict = dict(zip(self.name_list, self.threshold_cols))
        self.name_dict = dict(zip(self.name_list, self.id_list))

        if use_cust:
            # finalize excavation / terrain fill features
            self.fill_ids.append("cust")
            self.excavate_ids.append("cust")

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = FeatureDefinitions (%s)" % os.path.dirname(__file__))
        print(dir(self))


class ReachDefinitions:
    def __init__(self):
        self.internal_id = ["reach_00", "reach_01", "reach_02", "reach_03", "reach_04", "reach_05", "reach_06",
                            "reach_07", "none"]
        self.reach_no = range(0, self.internal_id.__len__())
        self.reader = cRM.Read()
        self.id_xlsx = self.reader.get_reach_info("id")
        self.id_dict = dict(zip(self.internal_id, self.id_xlsx))
        self.id_no_dict = dict(zip(self.id_xlsx, self.reach_no))

        self.names_xlsx = self.reader.get_reach_info("full_name")
        self.name_dict = dict(zip(self.internal_id, self.names_xlsx))

        # make dictionaries containing reach id information
        self.dict_id_names = {}
        self.dict_id_int_id = {}
        i = 0
        for id in self.id_xlsx:
            self.dict_id_names.update({id: self.names_xlsx[i]})
            self.dict_id_int_id.update({id: self.internal_id[i]})
            i += 1
        del i

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = ReachDefinitions (%s)" % os.path.dirname(__file__))
        print(dir(self))
