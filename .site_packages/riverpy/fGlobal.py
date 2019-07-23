try:
    import os, logging, sys, glob, webbrowser, time
    from collections import Iterable  # used in the flatten function
    from bisect import bisect_left
except:
    print("ExceptionERROR: Missing fundamental packages (required: bisect, collections, os, sys, glob, logging, time, webbrowser).")

try:
    import arcpy
    from arcpy.sa import *
except:
    logging.info("ArcGIS ERROR: No SpatialAnalyst extension available.")

try:
    import config
    import cDefinitions as cDef
except:
    print("ExceptionERROR: Cannot find package files (cDefinitions.py).")


def chk_is_empty(variable):
    try:
        value = float(variable)
    except ValueError:
        value = variable
        pass
    try:
        value = str(variable)
    except ValueError:
        pass
    return bool(value)


def chk_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def clean_dir(directory):
    # Delete everything reachable IN the directory named in 'directory',
    # assuming there are no symbolic links.
    # CAUTION:  This is dangerous!  For example, if directory == '/', it
    # could delete all your disk files.
    for root, dirs, files in os.walk(directory, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))


def cool_down(seconds):
    # Pauses script execution for the input argument number of seconds
    # seconds = INT
    sys.stdout.write('Cooling down (waiting for processes to end) ... ')
    for i in range(seconds, 0, -1):
        sys.stdout.write(str(i) + ' ')
        sys.stdout.flush()
        time.sleep(1)
    sys.stdout.write('\n')


def del_ovr_files(directory):
    # directory must end with "\\" or "/"
    all_files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    for f in all_files:
        if ".ovr" in f:
            try:
                print("Attempting to remove old temporary files ...")
                os.remove(directory + f)
                print("Success.")
            except:
                pass


def dict_values2list(dv):
    # converts a 'dict_values' object into a 'list'
    # also works for 'dict_keys' to 'list' conversion
    out_list = []
    [out_list.append(item) for item in dv]
    return out_list


def dict2str(dictionary, **kwargs):
    # converts a dict to a STR expression - reutrn "{e: 1, f: 2, ...}" - used in arcpy-calculatefiel mgmt
    inverse_dict = False  # optional keyword arg: if true: dictionary keys and values will be inversed
    try:
        for k in kwargs.items():
            if "inverse_dict" in k[0]:
                inverse_dict = k[1]
    except:
        pass
    dict_str = "{"
    cc = 1
    for k, v in dictionary.items():
        skey = "\'%s\'" % k if type(k) == str else str(k)
        sval = "\'%s\'" % v if type(v) == str else str(v)
        if not inverse_dict:
            dict_str += "{0}: {1}".format(skey, sval)
        else:
            dict_str += "{1}: {0}".format(skey, sval)
        if not (cc == dictionary.__len__()):
            dict_str += ", "
        else:
            dict_str += "}"
        cc += 1
    return dict_str


def eliminate_nan_from_list(base_list, *args):
    # eliminates nan values from a list and all other lists provided with *args
    partner_lists = []
    try:
        for partner_list in args:
            partner_lists.append(partner_list)
    except:
        pass

    test_list = base_list
    for val in base_list:
        try:
            test = float(val)
        except:
            while val in base_list:
                rem_index = base_list.index(val)
                del base_list[rem_index]
                try:
                    for partner_list in partner_lists:
                        del partner_list[rem_index]
                except:
                    pass
    partner_lists.insert(0, base_list)
    return partner_lists


def err_info(func):
    def wrapper(*args, **kwargs):
        arcpy.gp.overwriteOutput = True
        logger = logging.getLogger("logfile")
        try:
            func(*args, **kwargs)
        except arcpy.ExecuteError:
            logger.info("ExecuteERROR: (arcpy).")
            logger.info(arcpy.GetMessages(2))
            arcpy.AddError(arcpy.GetMessages(2))
        except Exception as e:
            logger.info("ExceptionERROR: (arcpy).")
            logger.info(e.args[0])
            arcpy.AddError(e.args[0])
        except:
            logger.info("ERROR: (arcpy).")
            logger.info(arcpy.GetMessages())
    return wrapper


def file_names_in_dir(directory):
    # returns file names only (without directory)
    return [name for name in os.listdir(directory) if os.path.isfile(os.path.join(directory, name))]


def flatten(lis):
    for item in lis:
        if isinstance(item, Iterable) and not isinstance(item, str):
            for x in flatten(item):
                yield x
        else:
            yield item


def get_closest_val_in_list(usr_list, target_num):
    """ Returns closes value to target_num in a sorted usr_list
    if two numbers are equally close the smallest number is returned
    """
    pos = bisect_left(usr_list, target_num)
    if pos == 0:
        return usr_list[0]
    if pos == len(usr_list):
        return usr_list[-1]
    before = usr_list[pos - 1]
    after = usr_list[pos]
    if after - target_num < target_num - before:
       return after
    else:
       return before


def get_credits():
    c_file = open(config.dir2templates + "credits.txt", "r")
    credits_str = "\n".join(c_file.read().splitlines())
    c_file.close()
    return credits_str


def get_newest_output_folder(directory):
    # return that the newest folder in a directory
    lof = glob.glob(directory + "*\\")
    latest_dir = max(lof, key=os.path.getctime)
    return str(latest_dir)


def get_subdir_names(directory):
    return [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]


def interpolate_linear(x1, x2, y1, y2, xi):
    # returns linear interpolation yi of xi between two points 1 and 2
    return y1 + ((xi - x1) / (x2 - x1) * (y2 - y1))


def list_file_type_in_dir(directory, f_ending):
    # directory = full directory ending on "/" or "\\"
    # f_ending = STR, e.g., ".py"
    # returns a LIST of full file paths
    return glob.glob(directory + "*" + f_ending)


def make_output_dir(condition, reach_ids, habitat_analysis, relevant_feat_names):
    # sets / creates a raster/shp/mxd/pdf output dir as a function of provided condition + feature layer type
    # condition = STR or number (will be converted to str)
    # reach_id = LIST from MT/.templates/computation_extents.xlsx
    # habitat_analysis = BOOL
    # relevant_feat_names = LIST with entries from LD/.templates/threshold_values.xlsx
    features = cDef.FeatureDefinitions()
    feat_id_list = []
    [feat_id_list.append(features.name_dict[item]) for item in relevant_feat_names]
    feat_col_list = []
    [feat_col_list.append(features.col_name_dict[item]) for item in relevant_feat_names]

    if reach_ids.__len__() == 1:
        if not str(reach_ids[0]) == "none":
            reach_name = "_" + str(reach_ids[0])
        else:
            reach_name = ""
    else:
        reach_name = "_all"

    type_int_list = []
    for tt in feat_col_list:
        if tt in features.threshold_cols_framework:
            type_int_list.append(1)
        if tt in features.threshold_cols_plants:
            type_int_list.append(2)
        if tt in features.threshold_cols_toolbox:
            type_int_list.append(2)
        if tt in features.threshold_cols_complement:
            type_int_list.append(3)
    try:
        feat_lyr_type = int(sum(type_int_list)/type_int_list.__len__())
    except:
        feat_lyr_type = 0

    if not habitat_analysis:
        # define output directory as a function of layer type
        if feat_lyr_type > 0:
            for i in range(0, 9):
                test_folder = str(condition) + reach_name + "_lyr" + str(feat_lyr_type) + str(i)
                test_dir = config.dir2lf + "Output\\Rasters\\" + test_folder + "\\"
                if not os.path.exists(test_dir):
                    os.makedirs(test_dir)
                    output_dir = test_dir
                    break
                else:
                    files_exist = True  # basic assumption: a raster for the analyzed feature already exists
                    file_count = len(
                        [name for name in os.listdir(test_dir) if os.path.isfile(os.path.join(test_dir, name))])
                    if file_count < 9:
                        output_dir = test_dir
                        break
                    else:
                        file_names = []
                        [file_names.append(fn) for fn in os.listdir(test_dir) if os.path.isfile(os.path.join(test_dir, fn))]
                        for f_id in feat_id_list:
                            for f_n in file_names:
                                if not(f_id in f_n):
                                    files_exist = False  # if any analyzed feature is not yet present, this folder is OK
                                    output_dir = test_dir
                                    break

                        if not files_exist:
                            break
                        if not (i < 9):
                            print("Maximum folder size for this layer reached -- restarting at lyrX0.")
                            print("Consider better file structure; this time, old files are deleted.")
                            test_folder = str(condition) + "_lyr" + str(feat_lyr_type) + str(0)
                            output_dir = config.dir2lf + "Output\\Rasters\\" + test_folder + "\\"
                            break
        else:
            output_dir = config.dir2lf + "Output\\Rasters\\" + str(condition) + reach_name + "lyr00\\"
    else:
        output_dir = config.dir2lf + "Output\\Rasters\\" + str(condition) + reach_name + "_hab\\"

    if not("output_dir" in locals()):
        print("No reach or feature layer or habitat_analysis information.")
        print("--> Output folder name corresponds to input condition.")
        output_dir = config.dir2lf + "Output\\Rasters\\" + str(condition) + "\\"

    chk_dir(output_dir)

    return output_dir


def open_file(full_file_path):
    _f = full_file_path
    if os.path.isfile(_f):
        try:
            webbrowser.open(_f)
            msg = ""
        except:
            msg = "Cannot open " + str(_f) + ". Ensure that your OS has a defined standard application for relevant file types (e.g., .inp or .xlsx)."
    else:
        msg = "The file \'\n" + str(_f) + "\'\n does not exist. Check MaxLifespan directory."
    return msg


def open_folder(directory):
    try:
        import subprocess
        # other python versions than 2.7: import subprocess32
        my_platform = sys.platform
        if my_platform[0:3].lower() == "win":
            # print("Hello Windows!")
            call_target = "explorer " + directory
            subprocess.call(call_target, shell=True)
            print("Found subprocess --> opening target folder.")
        if my_platform[0:3].lower() == "lin":
            # print("Hello Linux!")
            subprocess.check_call(['xdg-open', '--', directory])
            print("Found subprocess --> opening target folder.")
        if my_platform[0:3].lower() == "dar":
            # print("Hello Mac OS!")
            subprocess.check_call(['open', '--', directory])
            print("Found subprocess --> opening target folder.")
            try:
                os.system("start \"\" https://en.wikipedia.org/wiki/Criticism_of_Apple_Inc.")
            except:
                pass
    except:
        pass


def print_dict(dictionary):
    out_str = ""
    for k, v in dictionary.items():
        out_str += "   {0} - {1}".format(str(k), str(" + ".join(v)))
    return out_str


def raster2shp(raster_name, out_shp_name=str(), simplify="NO_SIMPLIFY", calculate_area=True):
    # raster_name = STR of full path to INTEGER Raster
    if out_shp_name.__len__() < 1:
        out_shp_name = raster_name.split(".")[0] + ".shp"
    arcpy.CheckOutExtension('Spatial')
    arcpy.RasterToPolygon_conversion(Int(arcpy.Raster(raster_name)), out_shp_name, simplify)
    if calculate_area:
        arcpy.AddField_management(out_shp_name, "F_AREA", "FLOAT", 9)
        arcpy.CalculateGeometryAttributes_management(out_shp_name, geometry_property=[["F_AREA", "AREA"]],
                                                     area_unit=out_shp_name)
    arcpy.CheckInExtension('Spatial')
    return out_shp_name


def read_txt(file_name):
    """ returns numeric data of a comma delimited text file
    INPUT:  file = full path of text (or csv) file
    OUTPUT: data = LIST: [[col0_row0, col1_row0, ...], [col0_row1, col1_row1, ...], ...]
    original use: ProjectMaker/s21_plantings_stabilization
    """

    logger = logging.getLogger("logfile")
    logger.info(" >> Reading " + str(file_name))
    data = []
    if os.path.isfile(file_name):
        file_name = open(file_name)
        lines = file_name.readlines()
        header = lines[0].split(",")
        try:
            col_grid = header.index("gridcode")
            try:
                col_area = header.index("F_sAREA\n")
            except:
                try:
                    col_area = header.index("F_AREA")
                except:
                    col_area = header.index("F_AREA\n")
        except:
            col_grid = 0
        lines = lines[1:]  # remove header
        for l in lines:
            try:
                l = str(l).split(",")
                grid_val = float(l[col_grid])
                try:
                    area_val = float(l[col_area].split("\n")[0])
                except:
                    area_val = float(l[col_area])
            except:
                logger.info("    !! non-numeric data in text file (line " + str(l) + ").")
                grid_val = 0.0
                area_val = 0.0
            data.append([grid_val, area_val])
        logger.info("     --> File read OK.")
    else:
        logger.info("ERROR: File not found.")
    return data


def rm_dir(directory):
    # Deletes everything reachable from the directory named in 'directory', and the directory itself
    # assuming there are no symbolic links.
    # CAUTION:  This is dangerous!  For example, if directory == '/' deletes all disk files
    for root, dirs, files in os.walk(directory, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))
    os.rmdir(directory)


def rm_file(full_name):
    # fullname = str of directory + file name
    try:
        os.remove(full_name)
    except:
        pass


def rm_raster(full_path):
    # Deletes everything reachable from the rasters full_path
    # full_path = "D:\\example_folder\\example_raster_name.tif"
    try:
        os.remove(full_path)
        return False
    except:
        locked = True
    try:
        if not locked:
            os.remove(full_path + ".aux.xml")
            return False
    except:
        locked = True
    try:
        if not locked:
            os.remove(full_path + ".ovr")
            return False
    except:
        locked = True
    try:
        if not locked:
            rm_dir(full_path + "\\")
            return False
    except:
        locked = True
    return locked


def shp2raster(shp_name, out_raster_name=str(), field_name=str(), cellsize=1):
    # shp_name = STR of full path to POLYGON
    # out_raster_name = STR to fill path of output (TIF) Raster
    if out_raster_name.__len__() < 1 or not(out_raster_name.endswith(".tif")):
        out_raster_name = shp_name.split(".")[0] + ".tif"
    arcpy.CheckOutExtension('Spatial')
    if field_name.__len__() < 1:
        fields = arcpy.ListFields(shp_name)
        if fields.__len__() > 2:
            for f in fields:
                if "gridcode" in f.name:
                    field_name = f.name
            if field_name.__len__() < 1:
                field_name = fields[2].name
        else:
            field_name = fields[0].name  # this is FID
    try:
        arcpy.PolygonToRaster_conversion(shp_name, value_field=field_name, out_rasterdataset=out_raster_name,
                                         cell_assignment="CELL_CENTER", priority_field=field_name, cellsize=cellsize)
    except arcpy.ExecuteError:
        logging.info("ExecuteERROR: (arcpy).")
        logging.info(arcpy.GetMessages(2))
        arcpy.AddError(arcpy.GetMessages(2))
    except Exception as e:
        logging.info("ExceptionERROR: (arcpy).")
        logging.info(e.args[0])
        arcpy.AddError(e.args[0])
    except:
        logging.info("ERROR: (arcpy).")
        logging.info(arcpy.GetMessages())
    arcpy.CheckInExtension('Spatial')
    return out_raster_name


def spatial_license(func):
    def wrapper(*args, **kwargs):
        arcpy.CheckOutExtension('Spatial')
        func(*args, **kwargs)
        arcpy.CheckInExtension('Spatial')
    return wrapper


def str2frac(arg):
    arg = arg.split('/')
    return int(arg[0]) / int(arg[1])


def str2num(arg, sep):
    # function converts string of type 'X[sep]Y' to number
    # sep is either ',' or '.'
    # e.g. '2,30' is converted with SEP = ',' to 2.3
    _num = arg.split(sep)
    _a = int(_num[0])
    _b = int(_num[1])
    _num = _a + _b * 10 ** (-1 * len(str(_b)))
    return _num


def str2tuple(arg):
    try:
        arg = arg.split(',')
    except ValueError:
        print('ERROR: Bad assignment of separator.\nSeparator must be [,].')
    tup = (int(arg[0]), int(arg[1]))
    return tup


def tuple2num(arg):
    # function converts float number with ',' separator for digits to '.' separator
    # type(arg) = tuple with two entries, e.g. (2,40)
    # call: tuple2num((2,3))
    new = arg[0] + arg[1] * 10 ** (-1 * len(str(arg[1])))
    return new


def verify_shp_file(shapefile):
    # shapefile = STR (full path) of shapefile
    # returns TRUE if the shapefile is NOT EMPTY (has at least one polygon or line)
    try:
        item_number = arcpy.GetCount_management(shapefile)
    except:
        return False
    if item_number > 0:
        return True
    else:
        return False


def write_data(folder_dir, file_name, data):
    if not os.path.exists(folder_dir):
        os.mkdir(folder_dir)
    os.chdir(folder_dir)

    f = open(file_name+'.txt', 'w')
    for i in data:
        line = str(i)+'\n'
        f.write(line)
    print('Data written to: \n' + folder_dir + '\\' + str(file_name) + '.csv')


def write_dict2xlsx(data_dict, file, key_col, val_col, start_row):
    """ uses openpyxl to write data to an xlsx-workbook
    INPUT:  file = full path of xlsx file
    """
    logger = logging.getLogger("logfile")
    try:
        # load relevant files from RiverArchitect/ModifyTerrain module
        sys.path.append(config.dir2oxl)
        import openpyxl as oxl
    except:
        logger.info("ExceptionERROR: Cannot find %s." % config.dir2oxl)
        logger.info("                --> Correction required in: fFunctions.py")
        try:
            from inspect import currentframe, getframeinfo
            frameinfo = getframeinfo(currentframe())
            logger.info("                    " + str(frameinfo.filename))
            logger.info("                    line number: " + str(frameinfo.lineno - 9))
        except:
            pass
    logger.info(" >> Writing to " + str(file))

    try:
        wb = oxl.load_workbook(filename=file)
    except:
        logger.info("ERROR: Failed to access " + str(file))
        return -1
    try:
        ws = wb["from_geodata"]
    except:
        logger.info("ERROR: Sheet \'from_geodata\' is missing in " + str(file))
        return -1

    i_row = start_row
    for dct_key in data_dict.keys():
        ws[key_col + str(i_row)].value = dct_key
        ws[val_col + str(i_row)].value = data_dict[dct_key]
        i_row += 1

    wb.save(str(file))
    wb.close()
    logger.info(" -- OK (Write to workbook)")

