#!/usr/bin/python
try:
    import os, sys, logging, shutil
    # load other RA routines
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\riverpy\\")
    import config
    sys.path.append(config.dir2oxl)
    import fGlobal as fGl
    import openpyxl as oxl  # modified package
    import datetime, random
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")


class MakeFlowTable:
    def __init__(self, condition, purpose, **kwargs):
        # condition = STR defining the CONDITION
        # purpose =  STR, either "sharc" or "q_return"  or "q_duration"
        # **kwargs: unit
        self.unit = ""
        # parse optional arguments
        try:
            for k in kwargs.items():
                if "unit" in k[0]:
                    self.unit = k[1]  # STR of either "us" or "si"
        except:
            pass
        self.out_ras = []
        self.condition = condition
        self.dict_Q_h_ras = {}
        self.dict_Q_u_ras = {}
        self.dict_Q_va_ras = {}

        self.dir_in_ras = config.dir2conditions + str(self.condition) + "\\"
        self.dir_xlsx_out = ""

        self.logger = logging.getLogger("logfile")
        self.discharges = []
        self.h_rasters = []
        self.u_rasters = []
        self.va_rasters = []

        # dummy workbook variable instantiations
        self.xlsx_template = config.xlsx_dummy
        self.wb = oxl.load_workbook(filename=self.xlsx_template)
        self.wb_out_name = ""
        self.ws = self.wb.worksheets[0]

        # workbook range specifications
        self.col_Q = ""
        self.col_ras_h = ""
        self.col_ras_u = ""
        self.row_start = int()
        self.set_directories(purpose)  # sets purpose-specific workbook ranges
        self.get_condition_discharges()
        try:
            fGl.chk_dir(self.dir_xlsx_out)
        except:
            pass

    def copy_wb(self, full_wb_inpath, full_save_path):
        try:
            self.logger.info("   * Creating a workbook copy: ")
            wb = oxl.load_workbook(filename=full_wb_inpath)
            self.logger.info("     " + full_save_path)
            wb.save(full_save_path)
            wb.close()
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def get_condition_discharges(self):
        ras_name_list = [rn for rn in os.listdir(self.dir_in_ras) if os.path.isdir(os.path.join(self.dir_in_ras, rn))]
        if ras_name_list.__len__() < 1:
            # look for geoTIFFs if ras_name_list is empty
            ras_name_list = [i for i in os.listdir(self.dir_in_ras) if i.endswith('.tif')]

        self.logger.info("   * analyzing relevant discharges and matching Rasters ...")
        for rn in ras_name_list:
            if rn[0] == "h":
                self.logger.info("     -- Found flow depth raster: " + str(rn))
                try:
                    _Q_ = fGl.read_Q_str(rn, prefix='h')
                    self.discharges.append(_Q_)
                except:
                    self.logger.info("ERROR: The Raster name is not coherent with the name conventions. Name correction needed.")
                self.h_rasters.append(rn)
            if rn[0] == "u":
                self.logger.info("     -- Found flow velocity raster: " + str(rn))
                self.u_rasters.append(rn)
            if rn[:2] == "va":
                self.logger.info("     -- Found flow velocity angle raster: " + str(rn))
                self.va_rasters.append(rn)

        # make Q_flowdur-raster dictionary
        self.dict_Q_h_ras = dict(zip(self.discharges, self.h_rasters))
        self.dict_Q_u_ras = dict(zip(self.discharges, self.u_rasters))
        self.dict_Q_va_ras = dict(zip(self.discharges, self.va_rasters))
        # order discharges in descending order for flow duration curve
        self.discharges.sort(reverse=True)

    def make_aquatic_condition_xlsx(self, fish_sn):
        # fish_sn == STR -- 4 digits indicating fish species and lifestage
        self.logger.info("   * using workbook template: " + str(self.xlsx_template).split("\\")[-1])
        # open relevant workbook
        self.wb_out_name = self.dir_xlsx_out + "{0}_sharea_{1}.xlsx".format(str(self.condition), fish_sn)
        self.open_wb_work_copy()
        self.write_flow_info2xlsx()
        return self.wb_out_name

    def make_condition_xlsx(self):
        self.logger.info("   * using workbook template: " + str(self.xlsx_template).split("\\")[-1])
        # open relevant workbook
        self.wb_out_name = self.dir_in_ras + "flow_definitions.xlsx"
        self.open_wb_work_copy()
        self.write_flow_info2xlsx()
        self.save_close_wb()
        return self.wb_out_name

    def open_wb(self, full_wb_path, ws):
        # full_wb_path = STR -- e.g., C:\\...\\my.xlsx
        # ws = STR or INT -- worksheet name or number in workbook
        self.logger.info("   * opening " + str(full_wb_path))
        try:
            self.wb = oxl.load_workbook(filename=full_wb_path)
        except:
            self.logger.info("ERROR: Failed to access " + str(full_wb_path).split("\\")[-1] + ".")
        try:
            ws = int(ws)
        except:
            pass
        if type(ws) == int:
            try:
                self.ws = self.wb.worksheets[ws]
            except:
                pass
        else:
            try:
                self.ws = self.wb.active
            except:
                self.logger.info("ERROR: Could not find any worksheet.")
                return -1

    def open_wb_work_copy(self):
        if os.path.isfile(self.wb_out_name):
            try:
                self.logger.info("   * creating safety copy of existing workbook ...")
                rnd_ext = str(random.randint(1000000, 9999999))
                shutil.copy2(self.wb_out_name, self.wb_out_name.split(".xlsx")[0] + "_old" + rnd_ext + ".xlsx")
                fGl.rm_file(self.wb_out_name)
            except:
                self.logger.info("ERROR: Failed to access " + str(self.wb_out_name) + ".")
                return -1
        # else:
        try:
            self.logger.info("   * Creating new workbook (copy from templates folder) ...")
            __wb__ = oxl.load_workbook(filename=self.xlsx_template)
        except:
            self.logger.info("ERROR: Failed to load " + str(self.xlsx_template) + ".")
            return -1
        self.wb = __wb__
        __wb__.close()
        try:
            self.ws = self.wb[str(self.wb.sheetnames[0])]
        except:
            self.logger.info("ERROR: TEMPLATE sheet does not exist.")
            return -1

    def read_float_column_short(self, column, start_read_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        # returns dictionary with cell data and cell information

        data = {}
        valid_content = True
        __row__ = start_read_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                valid_content = False
            if valid_content:
                data.update({cell_value: str(column) + str(__row__)})
                __row__ += 1
        self.logger.info("   * OK")
        return data

    def save_close_wb(self, *full_save_path):
        try:
            self.wb_out_name = full_save_path[0]
        except:
            pass

        try:
            self.logger.info("   * Saving as: " + self.wb_out_name)
            self.wb.save(self.wb_out_name)
            self.wb.close()
            self.logger.info("   * OK")
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def set_directories(self, purpose):
        # purpose =  STR, either "sharc" or "q_frequence"

        if purpose.lower() == "sharc":
            self.logger.info("   * scan %s for ecohydraulically relevant flows" % self.condition)
            self.xlsx_template = "{0}.templates\\CONDITION_sharea_template_{1}.xlsx".format(config.dir2sh, str(self.unit))
            self.dir_xlsx_out = config.dir2sh + "SHArea\\"
            self.col_Q = "B"
            self.col_ras_h = "C"
            self.col_ras_u = "D"
            self.row_start = 4

        if purpose.lower() == "q_return":
            self.logger.info("   * scan %s for available flows" % self.condition)
            self.xlsx_template = config.dir2flows + "templates\\flow_return_period_template.xlsx"
            self.dir_xlsx_out = self.dir_in_ras
            self.col_Q = "B"
            self.col_ras_h = "D"
            self.col_ras_u = "E"
            self.row_start = 5

    def write_data_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * ERROR: Could not write value to CELL " + str(column) + str(row))

    def write_data_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list must be a LIST object
        self.logger.info("   * Writing column data (starting at " + str(column) + str(start_row) + ") ...")
        __row__ = start_row
        for val in data_list:
            self.ws[str(column) + str(__row__)].value = val
            __row__ += 1
        self.logger.info("   * OK")

    def write_flow_info2xlsx(self):
        # write discharges and associated raster names
        self.logger.info("   * Writing discharges and associated raster names to workbook.")
        for iq in range(0, self.discharges.__len__()):
            try:
                self.ws[self.col_Q + str(self.row_start + iq)].value = self.discharges[iq]
                self.ws[self.col_ras_h + str(self.row_start + iq)].value = self.dict_Q_h_ras[self.discharges[iq]]
                self.ws[self.col_ras_u + str(self.row_start + iq)].value = self.dict_Q_u_ras[self.discharges[iq]]
            except:
                self.logger.info("ERROR: Invalid cell assignment for discharge / Rasters.")
        self.logger.info("   * OK")

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = MakeTable (%s)" % os.path.dirname(__file__))
        print(dir(self))
