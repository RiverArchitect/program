#!/usr/bin/python
# Filename: cReachManager.py
try:
    import os, sys, logging
    import datetime as dt
    sys.path.append(os.path.dirname(__file__))
    # import own functions -- make sure that all *.py files are in the same folder
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\openpyxl\\")
    import openpyxl as oxl  # modified package
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")


class Read:
    def __init__(self):
        # type defines lines to read in .inp file
        self.path2mt = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\..')) + "\\ModifyTerrain\\"
        self.xlsx_coord = self.path2mt + ".templates\\computation_extents.xlsx"
        self.logger = logging.getLogger("logfile")

        # define columns
        self.col_min_x = "D"
        self.col_max_x = "E"
        self.col_min_y = "F"
        self.col_max_y = "G"

        # define rows
        self.row_dict = {"reach_00": "6", "reach_01": "7", "reach_02": "8", "reach_03": "9", "reach_04": "10",
                         "reach_05": "11", "reach_06": "12", "reach_07": "13", "none": "none"}
        self.row_start = 6

    def get_reach_coordinates(self, internal_reach_id):
        self.logger.info(" ->> Reading Reach coordinates from computation_extents.xlsx ...")

        try:
            wb = oxl.load_workbook(filename=self.xlsx_coord, read_only=True, data_only=True)
        except:
            wb = ""
            self.logger.info("ERROR: Failed to access computation_extents.xlsx.")
        if internal_reach_id == "none":
            return "MAXOF"
        try:
            ws = wb['extents']
        except:
            ws = []
            self.logger.info("ERROR: Could not find sheet \'extents\' in computation_extents.xlsx.")
        try:
            min_x = float(ws[self.col_min_x + self.row_dict[internal_reach_id]].value)
            max_x = float(ws[self.col_max_x + self.row_dict[internal_reach_id]].value)
            min_y = float(ws[self.col_min_y + self.row_dict[internal_reach_id]].value)
            max_y = float(ws[self.col_max_y + self.row_dict[internal_reach_id]].value)
            self.logger.info(" ->> Success.")
            wb.close()
            return [min_x, min_y, max_x, max_y]
        except:
            self.logger.info("ERROR: Failed to read coordinates from computation_extents.xlsx (return 0).")
            return [0, 0, 0, 0]

    def get_reach_info(self, type):
        # type = full_name sets column read to "B"
        # type = id sets column read to "C"
        allowed_types = ["full_name", "id"]
        if type in allowed_types:
            if type == "full_name":
                col = "B"
                last_item = "Raster extents"
                self.logger.info(" ->> Reading Reach names (col. " + str(col) + ") from computation_extents.xlsx ...")
            if type == "id":
                col = "C"
                last_item = "none"
                self.logger.info(" ->> Reading Reach IDs (col. " + str(col) + ") from computation_extents.xlsx ...")
        else:
            last_item = "none"
            self.logger.info("WARNING: Invalid type assignment -- setting reach names to IDs.")
            col = "C"
        reach_list = []

        try:
            wb = oxl.load_workbook(filename=self.xlsx_coord, read_only=True)
            ws = wb['extents']

            cell_content = "init_string"
            i = self.row_start
            while not(cell_content.__len__() == 4):
                # __len__() == 4 corresponds to empty cells
                cell_content = str(ws[col + str(i)].value)
                if not(cell_content.__len__() == 4):
                    reach_list.append(cell_content)
                i += 1
                if i > 15:
                    self.logger.info("WARNING: computation_extents.xls contains too many reach names.")
                    break
            wb.close()
        except:
            self.logger.info("ERROR: failed to access computation_extents.xlsx (return empty list).")
        reach_list.append(last_item)  # append none-type = Raster extents
        return reach_list

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Read (%s)" % os.path.dirname(__file__))
        print(dir(self))


class Write:
    def __init__(self, output_dir):
        # ozutput_dir = STR of directory for output files (must end with \\ )
        self.xlsx_output_dir = output_dir
        self.vol_xlsx_template = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\..')) + "\\VolumeAssessment\\.templates\\volumes_template.xlsx"
        self.logger = logging.getLogger("logfile")

        # define columns
        self.col_reach = "B"
        self.col_volume = "C"

        # define rows
        self.row_start = 8

    def write_volumes(self, vol_name, reach_names, volumes, unit, vol_signature):
        # type(vol_name) ==  string
        # type(reach_names) ==  list of strings of full reach names
        # type(volumes) ==  list of floats (nested)
        # type(unit) ==  string (either cy or m3)
        # type(vol_signature) == INT: -1 --> negative volumes (ecav), +1 --> positive volume (fill)
        wb_name = str(vol_name) + "_volumes.xlsx"
        try:
            if os.path.isfile(self.xlsx_output_dir + wb_name):
                try:
                    self.logger.info("   * using " + str(self.xlsx_output_dir + wb_name) + " ...")
                    wb_new = oxl.load_workbook(filename=self.xlsx_output_dir + wb_name)
                except:
                    self.logger.info("ERROR: Failed to access " + str(self.xlsx_output_dir) + str(wb_name) + ".")
                    return -1
            else:
                try:
                    wb = oxl.load_workbook(filename=self.vol_xlsx_template)
                except:
                    self.logger.info("ERROR: Failed to access " + str(self.vol_xlsx_template) + ".")
                    return -1
                wb_new = wb
                wb.close()
            try:
                ws_temp = wb_new['template']
            except:
                self.logger.info("ERROR: TEMPLATE sheet does not exist.")
                return -1
            ws_new = wb_new.copy_worksheet(ws_temp)
            now = dt.datetime.now()
            if vol_signature < 0:
                prefix = "excavate_"
            else:
                prefix = "fill_"

            ws_new.title = prefix + '%04d' % (now.year,) + '%02d' % (now.month,) + '%02d' % (
                            now.day,) + "_" + '%02d' % (now.hour,) + "h" + '%02d' % (now.minute,)

            # clear unnecessary template cells
            try:
                ws_new["B1"] = ""
                ws_new["B1"]._style = ws_new["A1"]._style
                ws_new["C1"]._style = ws_new["A1"]._style
            except:
                self.logger.info("WARNING: Could not reset styles.")

            # write vol_name
            ws_new["C2"] = str(vol_name)
            # write volume signature
            if vol_signature < 0:
                ws_new["C3"] = "Excavation"
            else:
                ws_new["C3"] = "Fill"

            # write full feature names
            try:
                ws_new[chr(67) + str(self.row_start-3)] = vol_name
                ws_new[chr(67) + str(self.row_start-3)]._style = ws_new[self.col_volume + str(self.row_start-3)]._style
            except:
                self.logger.info("ERROR: Invalid volume name.")

            # write reach names
            i = self.row_start
            for rn in reach_names:
                ws_new[self.col_reach + str(i)] = rn
                i += 1
            del i

            # write volumes
            try:
                ws_new[self.col_volume + str(self.row_start - 1)] = "(" + str(unit) + ")"
            except:
                self.logger.info("WARNING: Failed to write unit system to worksheet.")

            j = ord(self.col_volume)  # convert char('C') to ascii value
            self.logger.info("   * writing %s volumes ... " % prefix.strip("_"))
            i = self.row_start
            for vol in volumes:
                try:
                    val = float(vol)
                except:
                    val = 0.0
                try:
                    ws_new[chr(j) + str(i)] = val
                except:
                    self.logger.info("ERROR: Volume value assignment failed.")
                i += 1  # increase row counter
            del i

            # sort sheets (bring latest to front)
            try:
                wb_new._sheets.sort(key=lambda ws: ws.title)
            except:
                self.logger.info("WARNING: Failed to arrange worksheets.")

            # save and close workbook: once saved, the same wb cannot be saved with the same name
            self.logger.info("   * saving " + str(self.xlsx_output_dir + wb_name) + "(new sheet name: " + str(ws_new.title) + ") ...")
            try:
                wb_new.save(self.xlsx_output_dir + wb_name)
                self.logger.info("FINISHED.")
            except:
                self.logger.info("ERROR: Writing failed.")
            wb_new.close()

        except:
            self.logger.info("ERROR: Failed to create " + str(self.xlsx_output_dir + wb_name) + ".")

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (%s)" % os.path.dirname(__file__))
        print(dir(self))
