#!/usr/bin/python
# Filename: cTerrainIO.py
try:
    import os, sys, logging
    import datetime as dt
    # import own functions -- make sure that all *.py files are in the same folder
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\openpyxl\\")
    import openpyxl as oxl  # modified package!
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")


class IOlogger:
    def __init__(self, logfile_name):
        self.logger = self.logging_start(logfile_name)

    def logging_start(self, logfile_name):
        logfilenames = ["error.log", logfile_name + ".log", "logfile.log"]
        for fn in logfilenames:
            fn_full = os.path.join(os.getcwd(), fn)
            if os.path.isfile(fn_full):
                try:
                    os.remove(fn_full)
                except:
                    pass
        # start logging
        logger = logging.getLogger(logfile_name)
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter("%(asctime)s - %(message)s")

        # create console handler and set level to info
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        # create error file handler and set level to error
        err_handler = logging.FileHandler(os.path.join(os.getcwd(), logfilenames[0]), "w", encoding=None, delay="true")
        err_handler.setLevel(logging.ERROR)
        err_handler.setFormatter(formatter)
        logger.addHandler(err_handler)
        # create debug file handler and set level to debug
        debug_handler = logging.FileHandler(os.path.join(os.getcwd(), logfilenames[1]), "w")
        debug_handler.setLevel(logging.DEBUG)
        debug_handler.setFormatter(formatter)
        logger.addHandler(debug_handler)
        return logger

    def logging_stop(self):
        # stop logging and release logfile
        for handler in self.logger.handlers:
            handler.close()
            self.logger.removeHandler(handler)

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = IOlogger (Module: ModifyTerrain)")


class Read:
    def __init__(self):
        # type defines lines to read in .inp file
        self.path2mt = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\..')) + "\\ModifyTerrain\\"
        self.xlsx_coord = self.path2mt + ".templates\\computation_extents.xlsx"

        # define columns
        self.col_min_x = "D"
        self.col_max_x = "E"
        self.col_min_y = "F"
        self.col_max_y = "G"

        # define rows
        self.row_dict = {"reach_00": "6", "reach_01": "7", "reach_02": "8", "reach_03": "9", "reach_04": "10",
                         "reach_05": "11", "reach_06": "12", "reach_07": "13"}
        self.row_start = 6

    def get_reach_coordinates(self, internal_reach_id):
        logger = IOlogger("IO_logger")
        logger.logger.info(" ->> Reading Reach coordinates from computation_extents.xlsx ...")

        try:
            wb = oxl.load_workbook(filename=self.xlsx_coord, read_only=True, data_only=True)
        except:
            wb = ""
            logger.logger.info("ERROR: Failed to access computation_extents.xlsx.")
        try:
            ws = wb['extents']
        except:
            ws = []
            logger.logger.info("ERROR: Could not find sheet \'extents\' in computation_extents.xlsx.")
        try:
            min_x = float(ws[self.col_min_x + self.row_dict[internal_reach_id]].value)
            max_x = float(ws[self.col_max_x + self.row_dict[internal_reach_id]].value)
            min_y = float(ws[self.col_min_y + self.row_dict[internal_reach_id]].value)
            max_y = float(ws[self.col_max_y + self.row_dict[internal_reach_id]].value)
            logger.logger.info(" ->> Success.")
            wb.close()
            logger.logging_stop()
            return [min_x, min_y, max_x, max_y]
        except:
            logger.logger.info("ERROR: Failed to read coordinates from computation_extents.xlsx (return 0).")
            logger.logging_stop()
            return [0, 0, 0, 0]

    def get_reach_info(self, type):
        # type = full_name sets column read to "B"
        # type = id sets column read to "C"
        logger = IOlogger("IO_logger")
        allowed_types = ["full_name", "id"]
        if type in allowed_types:
            if type == "full_name":
                col = "B"
                logger.logger.info(" ->> Reading Reach names (col. " + str(col) + ") from computation_extents.xlsx ...")
            if type == "id":
                col = "C"
                logger.logger.info(" ->> Reading Reach IDs (col. " + str(col) + ") from computation_extents.xlsx ...")
        else:
            logger.logger.info("WARNING: Invalid type assignment -- setting reach names to IDs.")
            col = "C"
        reach_list = []

        try:
            wb = oxl.load_workbook(filename=self.xlsx_coord, read_only=True)
            ws = wb['extents']

            cell_content = "init_string"
            i = self.row_start
            while not(cell_content.__len__() == 4):
                # __len__() == 4 corresponds to empty cells
                cell_content = str(ws[col + str(i)].value)
                if not(cell_content.__len__() == 4):
                    reach_list.append(cell_content)
                i += 1
                if i > 15:
                    logger.logger.info("WARNING: computation_extents.xls contains too many reach names.")
                    break
            wb.close()
        except:
            logger.logger.info("ERROR: failed to access computation_extents.xlsx (return empty list).")
        logger.logging_stop()
        return(reach_list)

    def __call__(self):
        print("Class Info: <type> = Read (Module: ModifyTerrain)")


class Write:
    def __init__(self):
        self.path2mt = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\..')) + "\\ModifyTerrain\\"
        self.xlsx_output_dir = self.path2mt + "Output\\Workbooks\\"
        self.logger = logging.getLogger("IO_info")

        # define columns
        self.col_reach = "B"
        self.col_volume = "C"

        # define rows
        self.row_start = 8

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def write_volumes(self, condition, features, reach_names, volumes, unit, vol_signature):
        # type(condition) ==  string
        # type(features) ==  list of strings
        # type(reach_names) ==  list of strings of full reach names
        # type(volumes) ==  list of list of floats (nested)
        # type(unit) ==  string (either cy or m3)
        # type(vol_signature) == INT: -1 --> negative volumes (ecav), +1 --> positive volume (fill)
        logger = IOlogger("IO_logger")
        wb_name = str(condition) + "_volumes.xlsx"
        try:
            if os.path.isfile(self.xlsx_output_dir + str(wb_name)):
                try:
                    logger.logger.info(" ->> Opening ModifyTerrain/Output/Workbooks/" + str(wb_name) + " ...")
                    wb_new = oxl.load_workbook(filename=self.xlsx_output_dir + wb_name)
                except:
                    logger.logger.info("ERROR: Failed to access " + str(self.xlsx_output_dir) + str(wb_name) +".")
                    return -1
            else:
                try:
                    wb = oxl.load_workbook(filename=self.xlsx_output_dir + "volumes_template.xlsx")
                except:
                    logger.logger.info("ERROR: Failed to access " + str(self.xlsx_output_dir) + str(wb_name) + ".")
                    return -1
                wb_new = wb
                wb.close()
            try:
                ws_temp = wb_new['template']
            except:
                logger.logger.info("ERROR: TEMPLATE sheet does not exist.")
                return -1
            ws_new = wb_new.copy_worksheet(ws_temp)
            now = dt.datetime.now()
            if vol_signature < 0:
                prefix = "excavate_"
            else:
                prefix = "fill_"

            ws_new.title = prefix + '%04d' % (now.year,) + '%02d' % (now.month,) + '%02d' % (
                            now.day,) + "_" + '%02d' % (now.hour,) + "h" + '%02d' % (now.minute,)

            # clear unnecessary template cells
            # white_fill = oxl.styles.PatternFill(start_color=self.white, end_color=self.white, fill_type='solid')
            try:
                ws_new["B1"] = ""
                ws_new["B1"]._style = ws_new["A1"]._style
                ws_new["C1"]._style = ws_new["A1"]._style
            except:
                logger.logger.info("WARNING: Could not reset styles.")

            # write condition
            ws_new["C2"] = str(condition)
            # write volume signature
            if vol_signature < 0:
                ws_new["C3"] = "Excavation"
            else:
                ws_new["C3"] = "Fill"

            # write full feature names
            i = 67  # start ascii character number corresponds to "C"
            for fn in features:
                try:
                    ws_new[chr(i) + str(self.row_start-3)] = fn
                    ws_new[chr(i) + str(self.row_start-3)]._style = ws_new[self.col_volume + str(self.row_start-3)]._style
                except:
                    logger.logger.info("ERROR: Invalid feature names for column headers.")
                i += 1
            del i

            # write reach names
            i = self.row_start
            for rn in reach_names:
                ws_new[self.col_reach + str(i)] = rn
                i += 1
            del i

            # write volumes
            try:
                ws_new[self.col_volume + str(self.row_start - 1)] = "(" + str(unit) + ")"
            except:
                logger.logger.info("WARNING: Failed to write unit system to worksheet.")

            j = ord(self.col_volume)  # convert char('C') to ascii value
            for vol_list in volumes:
                i = self.row_start
                cum_sum = 0
                for vol in vol_list:
                    if i == self.row_start:
                        # make column header
                        try:
                            ws_new[chr(j) + str(i - 2)]._style = ws_new[self.col_volume + str(i - 2)]._style
                            ws_new[chr(j) + str(i - 1)]._style = ws_new[self.col_volume + str(i - 1)]._style
                            ws_new[chr(j) + str(i - 2)].value = ws_new[self.col_volume + str(i - 2)].value
                            ws_new[chr(j) + str(i - 1)].value = ws_new[self.col_volume + str(i - 1)].value
                        except:
                            logger.logger.info("WARNING: Could not write headers.")
                    try:
                        val = float(vol)
                    except:
                        val = 0.0
                    try:
                        ws_new[chr(j) + str(i)] = val
                    except:
                        logger.logger.info("ERROR: Volume value assignment failed.")
                    cum_sum += val

                    if j > ord(self.col_volume):
                        # update cell styles if more than one feature
                        ws_new[chr(j) + str(i)]._style = ws_new[self.col_volume + str(i)]._style
                    i += 1  # increase row counter
                # sum up
                ws_new[chr(j) + str(i)] = cum_sum
                ws_new[chr(j) + str(i)]._style = ws_new[self.col_volume + str(i)]._style
                j += 1  # increase column counter
            del i, j

            # sort sheets (bring latest to front)
            try:
                wb_new._sheets.sort(key=lambda ws: ws.title)
            except:
                logger.logger.info("WARNING: Failed to arrange worksheets.")

            # save and close workbook: once saved, the same wb cannot be saved with the same name
            logger.logger.info(" ->> Writing to ModifyTerrain/Output/Workbooks/" + str(wb_name) + "(sheet name:" + str(
                ws_new.title) + ") ...")
            try:
                wb_new.save(self.xlsx_output_dir + wb_name)
                logger.logger.info("FINISHED.")
            except:
                logger.logger.info("ERROR: Writing failed.")
            wb_new.close()

        except:
            logger.logger.info("ERROR: Failed to create " + str(wb_name) + ".")
        logger.logging_stop()

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (Module: ModifyTerrain)")

