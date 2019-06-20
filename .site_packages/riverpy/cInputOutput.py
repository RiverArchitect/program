#!/usr/bin/python
try:
    import os, sys, logging, datetime
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging, datetime).")
try:
    import config
    sys.path.append(config.dir2oxl)
    import openpyxl as oxl  # modified package
except:
    print("ERROR: Could not load openpyxl.")


class Read:
    def __init__(self, full_file_name, *args, **kwargs):
        # args[0] = BOOL that prevents opening the xlsx file in read_mode (FALSE -> WB is not OPENED)
        # kwargs: ws = INT
        # ** kwargs: worksheet_no
        self.ws_no = 0
        # parse optional arguments
        try:
            for k in kwargs.items():
                if "worksheet_no" in k[0]:
                    self.ws_no = k[1]
        except:
            pass

        self.header_dict = {}
        self.logger = logging.getLogger("logfile")
        self.ws_data = []
        self.xlsx_file = full_file_name
        try:
            if not args[0]:
                self.open_wb(full_file_name, self.ws_no)
        except:
            self.open_wb(full_file_name, self.ws_no)

    def close_wb(self):
        try:
            self.wb.close()
        except:
            pass

    def open_wb(self, full_file_name, ws, **kwargs):
        # full_file_name = full path to wb
        # ws = INT (or STR) of worksheet number (or name)
        # keyword arguments (kwargs:
        #   read_mode: BOOL that defines read_only - default = True
        #   data_only: BOOL that defines if only data should be read - default = True
        read_mode = True
        direct_data = True
        # parse keyword arguments
        try:
            for k in kwargs.items():
                if "read_mode" in k[0]:
                    read_mode = k[1]
                if "direct_data" in k[0]:
                    direct_data = k[1]
        except:
            pass
        try:
            self.wb = oxl.load_workbook(filename=full_file_name, read_only=read_mode, data_only=direct_data)
        except:
            self.wb = ""
            self.logger.info("ERROR: Failed to access " + str(full_file_name).split("\\")[-1] + ".")
            self.logger.info("       (full name: " + str(full_file_name) + ")")
        try:
            self.ws = self.wb.worksheets[ws]
        except:
            try:
                self.ws = self.wb.active
            except:
                self.logger.info("ERROR: Could not find any worksheet.")
                return -1

    def read_column(self, column, start_row):
        # ATTENTION: Empty cells in a ws are ignored, i.e., if col A=empty, then col B shifts to col A
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # column = STR, e.g., "A"
        # start_row = INT
        if not self.ws_data.__len__() > 0:
            try:
                self.ws_data = self.read_ws()
            except:
                self.logger.info("ERROR: Could not read source file.")
                return -1
        column_values = []
        column_header = self.header_dict[column]
        for i in range(start_row - 2, self.ws_data.__len__() + 1):
            try:
                column_values.append(self.ws_data[i][column_header])
            except:
                pass
        return column_values

    def read_date_from_cell(self, column, row):
        # column = CHR - cell column
        # row = INT - cell row
        # reads COLUMN / ROW cell
        try:
            cell_value = self.ws[str(column) + str(row)].value
        except:
            cell_value = "None"
            self.logger.info("WARNING: Undefined cell " + str(column) + str(row))

        if type(cell_value) == datetime.datetime:
            return {'year': int(cell_value.year), 'month': int(cell_value.month), 'day': int(cell_value.day)}
        else:
            self.logger.info("WARNING: Cell " + str(column) + str(row) + "has not a recognized date format.")
            return {cell_value}

    def read_float_column_short(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        self.logger.info(
            "   * reading data column from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(column) + str(
                start_row) + ") ...")

        data = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                try:
                    cell_value = self.ws[str(column) + str(__row__)].value
                    if not cell_value:
                        valid_content = False
                except:
                    valid_content = False
            if valid_content:
                data.append(cell_value)
                __row__ += 1
        return data

    def read_row_str(self, row, start_col, **kwargs):
        # reads ROW beginning at COL until it meets an empty cell
        # row = INT
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
                if "end_col" in opt_var[0]:
                    # last relevant column
                    end_col = opt_var[1]
                if "if_row" in opt_var[0]:
                    # row that tells if the content of the actual cell is relevant
                    if_row = opt_var[1]
        except:
            pass

        if not ("col_skip" in locals()):
            col_skip = 1
        if not ("end_col" in locals()):
            end_col = "XFA"
        if not ("if_row" in locals()):
            if_row = row

        str_data = []
        valid_content = True
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        while valid_content:
            cell_string = str(self.ws[str(__col__) + str(row)].value)
            valid_content = self.test_cell_content(str(__col__) + str(if_row))
            if valid_content:
                str_data.append(cell_string)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
            else:
                valid_content = False
        return str_data

    def read_cell(self, col, row):
        try:
            return float(self.ws[str(col) + str(row)].value)
        except:
            try:
                self.ws[str(col) + str(row)].value
            except:
                return -1

    def read_ws(self):
        rows = self.ws.rows
        first_row = [cell.value for cell in next(rows)]
        data = []
        for row in rows:
            record = {}
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[key] = cell.value.strip()
                else:
                    record[key] = cell.value
            data.append(record)

        # assign column head names
        col_no = 64 + self.ws.min_column
        for head in data[0].keys():
            self.header_dict.update({self.col_num_to_name(col_no): head})
            col_no += 1
        return data

    def test_cell_content(self, cell_str):
        # cell_str = str, e.g., "C4"
        if str(self.ws[cell_str].value) == "None":
            return False
        else:
            return True

    @staticmethod
    def col_name_to_num(letters):
        pow = 1
        col_int = 0
        for letter in letters[::-1]:
            col_int += (int(letter, 36) - 9) * pow
            pow *= 26
        return col_int + 64

    @staticmethod
    def col_num_to_name(col_int):
        col_int -= 64
        letters = ''
        while col_int:
            mod = (col_int - 1) % 26
            letters += chr(mod + 65)
            col_int = (col_int - 1) // 26
        return ''.join(reversed(letters))

    def __call__(self):
        print("Class Info: <type> = Read (%s)" % os.path.dirname(__file__))
        print(dir(self))


class Write(Read):
    def __init__(self, template, **kwargs):
        # ** kwargs: worksheet_no
        Read.__init__(self, template, False)
        # parse optional arguments
        try:
            for k in kwargs.items():
                if "worksheet_no" in k[0]:
                    self.ws_no = k[1]
        except:
            pass
        try:
            self.open_wb(str(template), self.ws_no, read_mode=False, direct_data=True)
        except:
            self.logger.info("ERROR: Could not open workbook (cInputOutput.Write.__init__() -- " + str(template) + ").")

        self.wb_out_name = ""

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def save_close_wb(self, *full_save_path):
        try:
            self.wb_out_name = full_save_path[0]
        except:
            pass

        try:
            self.logger.info("   * saving as: " + self.wb_out_name)
            self.wb.save(self.wb_out_name)
            self.wb.close()
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def write_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * WARNING: Could not write value to CELL " + str(column) + str(row))

    def write_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list must be a LIST object
        self.logger.info("   * writing column data (starting at " + str(column) + str(start_row) + ") ...")
        __row__ = start_row
        for val in data_list:
            self.ws[str(column) + str(__row__)].value = val
            __row__ += 1

    def write_matrix(self, start_col, start_row, data_matrix):
        # start_col = STR, e.g., "A"
        # start_row = INT, e.g., 5
        # data_matrix must be a NESTED LIST (LIST(LIST))
        __row__ = start_row
        for n in data_matrix:
            __col__ = start_col
            __col_num__ = self.col_name_to_num(__col__)
            for m in n:
                self.ws[str(__col__) + str(__row__)].value = m
                __col_num__ += 1
                __col__ = self.col_num_to_name(__col_num__)
            __row__ += 1

    def write_row(self, start_col, row, data_list, **kwargs):
        # write a data list to ROW beginning at COL
        # data_list = LIST
        # row = INT
        # start_col = CHR, e.g., start_col = "B"
        col_skip = 1
        # parse optional keyword arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
        except:
            pass

        __col__ = start_col
        for e in data_list:
            try:
                self.write_cell(__col__, row, float(e))
            except:
                self.logger.info("   * WARNING: Non-numeric data in cell " + str(__col__) + str(row))
                self.write_cell(__col__, row, e)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip
            __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (%s)" % os.path.dirname(__file__))
        print(dir(self))

