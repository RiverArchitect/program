#!/usr/bin/python
# Filename: cInputOutput.py
try:
    import os, sys, logging
    import datetime as dt
    # import own functions -- make sure that all *.py files are in the same folder
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "\\.openpyxl\\") # contains adapted openpyxl package
    import openpyxl as oxl  # modified package!
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")


class IOlogger:
    def __init__(self, logfile_name):
        self.logger = self.logging_start(logfile_name)

    def logging_start(self, logfile_name):
        logfilenames = ["error.log", logfile_name + ".log", "logfile.log"]
        for fn in logfilenames:
            fn_full = os.path.join(os.getcwd(), fn)
            if os.path.isfile(fn_full):
                try:
                    os.remove(fn_full)
                except:
                    pass
        # start logging
        logger = logging.getLogger(logfile_name)
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter("%(asctime)s - %(message)s")

        # create console handler and set level to info
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        # create error file handler and set level to error
        err_handler = logging.FileHandler(os.path.join(os.getcwd(), logfilenames[0]), "w", encoding=None, delay="true")
        err_handler.setLevel(logging.ERROR)
        err_handler.setFormatter(formatter)
        logger.addHandler(err_handler)
        # create debug file handler and set level to debug
        debug_handler = logging.FileHandler(os.path.join(os.getcwd(), logfilenames[1]), "w")
        debug_handler.setLevel(logging.DEBUG)
        debug_handler.setFormatter(formatter)
        logger.addHandler(debug_handler)
        return(logger)

    def logging_stop(self):
        # stop logging and release logfile
        for handler in self.logger.handlers:
            handler.close()
            self.logger.removeHandler(handler)


class Read:
    def __init__(self):

        # type defines lines to read in .inp file
        self.xlsx_input = os.path.dirname(os.path.abspath(__file__)) + "\\.templates\\computation_extents.xlsx"

        # define columns
        self.col = "D"

    def get_data(self, col, row):
        # col = capital letter (A, B, ...)
        # row = integer
        logger = IOlogger("IO_logger")
        logger.logger.info("Reading Reach coordinates from computation_extents.xlsx ...")

        try:
            wb = oxl.load_workbook(filename=self.xlsx_input, read_only=True, data_only=True)
        except:
            wb = ""
            logger.logger.info("ERROR: Failed to access input.xlsx.")
        try:
            ws = wb['input']
        except:
            ws = []
            logger.logger.info("ERROR: Could not find sheet \'extents\' in computation_extents.xlsx.")
        try:
            data_read = float(ws[str(col) + str(row)].value)
            logger.logger.info("Success.")
            wb.close()
            logger.logging_stop()
            return ([data_read])
        except:
            logger.logger.info("ERROR: Failed to read coordinates from computation_extents.xlsx (return 0).")
            logger.logging_stop()
            return [0]

    def __call__(self):
        print("Class Info: <type> = Read (cInputOutput)")


class Write:
    def __init__(self, wb_out_name):
        self.xlsx_template_dir = os.path.dirname(os.path.abspath(__file__)) + "\\.templates\\"
        self.xlsx_output_dir = os.path.dirname(os.path.abspath(__file__)) + "\\"
        self.logger = logging.getLogger("IO_logger")

        # define columns
        self.col = "B"
        self.wb_out_name = wb_out_name  # NO FULL PATH, name.xlsx only

        # define rows
        self.row_start = 8

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def open_wb(self, template_name, ws):
        # template_name = STR -- e.g.,my.xlsx, must be situated in .templates/ -subfolder
        # ws = STR or INT -- worksheet name or number in workbook
        self.logger.info("   * using " + self.xlsx_template_dir + template_name)
        self.wb = oxl.load_workbook(filename=self.xlsx_template_dir + template_name)
        try:
            ws = int(ws)
        except:
            pass
        if type(ws) == int:
            try:
                self.ws = self.wb[str(self.wb.sheetnames[ws])]
            except:
                pass
        else:
            self.ws = self.wb[ws]

    def save_close_wb(self):
        try:
            self.logger.info("   * saving xlsx as:" + self.xlsx_output_dir + self.wb_out_name)
            self.wb.save(self.wb_out_name)
            self.wb.close()
            self.logger.info("   * ok")
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def write_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * ERROR: Could not write value to CELL " + str(column) + str(row))

    def write_column(self, data_list, col, start_row):
        i_row = start_row
        for it in data_list:
            self.ws[col + str(i_row)].value = it
            i_row += 1

    def write_dict2xlsx(self, data_dict, key_col, val_col, start_row):
        i_row = start_row
        for dct_key in data_dict.keys():
            self.ws[key_col + str(i_row)].value = dct_key
            self.ws[val_col + str(i_row)].value = data_dict[dct_key]
            i_row += 1

    def write_pool_riffle(self, data_list):
        # contains: B_p, B_r, D_z, h_p, h_r, h_rev, l_pool2pool, l_pool2riffle, m_pool, m_riffle, Q_rev, m, n, w_b
        # type(B_p) ==  float
        # type(B_r) ==  float
        # type(D_z) ==  float
        # type(h_p) ==  float
        # type(h_r) ==  float
        # type(l_pool2pool) == float
        # type(l_pool2riffle) == float
        # type(z_p) == float
        logger = IOlogger("IO_logger")
        wb_template = "output.xlsx"
        try:
            if os.path.isfile(self.xlsx_template_dir + str(wb_template)):
                try:
                    logger.logger.info("Using " + str(wb_template) + " ...")
                    wb = oxl.load_workbook(filename=self.xlsx_template_dir + wb_template)
                    wb_new = wb
                    wb.close()
                except:
                    logger.logger.info("ERROR: Failed to access " + str(self.xlsx_template_dir) + str(wb_template) + ".")
                    return -1

            try:
                ws_temp = wb_new['.template_pr']
            except:
                logger.logger.info("ERROR: TEMPLATE sheet does not exist.")
                return -1
            ws_new = wb_new.copy_worksheet(ws_temp)
            now = dt.datetime.now()
            prefix = "pool_riffle_"
            ws_new.title = prefix + '%04d' % (now.year,) + '%02d' % (now.month,) + '%02d' % (
                            now.day,) + "_" + '%02d' % (now.hour,) + "h" + '%02d' % (now.minute,)

            # write data
            col = "D"   # all
            #            B_p, B_r, D_z, h_p, h_r, h_rev, lp2p, lp2r, m_p, m_r, Q_rev, m, n, w_base
            data_rows = [  7,  13,   9,  10,  15,    25,   18,   19,   8,  14,    24,23,26, 22]
            row_wise_data_dict = dict(zip(data_rows, data_list))

            for row in row_wise_data_dict.keys():
                try:
                    ws_new[col + str(row)].value = row_wise_data_dict[row]
                except:
                    logger.logger.info("ERROR: Invalid cell parameters.")

            # save and close workbook: once saved, the same wb cannot be saved with the same name
            logger.logger.info("Writing to " + str(self.xlsx_output_dir) + str(self.wb_out_name) + "(sheet name:" + str(
                ws_new.title) + ") ...")
            try:
                wb_new.save(self.xlsx_output_dir + self.wb_out_name)
                logger.logger.info("FINISHED.")
            except:
                logger.logger.info("ERROR: Writing failed.")
            wb_new.close()

        except:
            logger.logger.info("ERROR: Failed to create " + str(self.wb_out_name) + ".")
        logger.logging_stop()

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (cInputOutput)")




