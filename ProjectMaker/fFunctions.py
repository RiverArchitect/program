import os, logging, sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\riverpy\\")
import config

def chk_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def del_ovr_files(directory):
    # directory must end with "\\" or "/"
    all_files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    for f in all_files:
        if ".ovr" in f:
            try:
                print("Attempting to remove old temporary files ...")
                os.remove(directory + f)
                print("Success.")
            except:
                pass


def read_txt(file_name):
    # returns numeric data of a comma delimited text file
    # INPUT:  file = full path of text (or csv) file
    # OUTPUT: data = LIST: [[col0_row0, col1_row0, ...], [col0_row1, col1_row1, ...], ...]
    logger = logging.getLogger("logfile")
    logger.info(" >> Reading " + str(file_name))
    data = []
    if os.path.isfile(file_name):
        file_name = open(file_name)
        lines = file_name.readlines()
        header = lines[0].split(",")
        try:
            col_grid = header.index("gridcode")
            try:
                col_area = header.index("F_sAREA\n")
            except:
                try:
                    col_area = header.index("F_AREA")
                except:
                    col_area = header.index("F_AREA\n")
        except:
            col_grid = 0
        lines = lines[1:]  # remove header
        for l in lines:
            try:
                l = str(l).split(",")
                grid_val = float(l[col_grid])
                try:
                    area_val = float(l[col_area].split("\n")[0])
                except:
                    area_val = float(l[col_area])
            except:
                logger.info("    !! non-numeric data in text file (line " + str(l) + ").")
                grid_val = 0.0
                area_val = 0.0
            data.append([grid_val, area_val])
        logger.info("     --> File read OK.")
    else:
        logger.info("ERROR: File not found.")
    return data


def rm_dir(directory):
    # Deletes everything reachable from the directory named in 'directory', and the directory itself
    # assuming there are no symbolic links.
    # CAUTION:  This is dangerous!  For example, if directory == '/' deletes all disk files
    for root, dirs, files in os.walk(directory, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))
    os.rmdir(directory)


def write_dict2xlsx(data_dict, file, key_col, val_col, start_row, *logger):
    # uses openpyxl to write data to an xlsx-workbook
    # INPUT:  file = full path of xlsx file
    logger = logging.getLogger("logfile")
    try:
        # load relevant files from RiverArchitect/ModifyTerrain module
        sys.path.append(config.dir2oxl)
        import openpyxl as oxl
    except:
        logger.info("ExceptionERROR: Cannot find %s." % config.dir2oxl)
        logger.info("                --> Correction required in: fFunctions.py")
        try:
            from inspect import currentframe, getframeinfo
            frameinfo = getframeinfo(currentframe())
            logger.info("                    " + str(frameinfo.filename))
            logger.info("                    line number: " + str(frameinfo.lineno - 9))
        except:
            pass
    logger.info(" >> Writing to " + str(file))

    try:
        wb = oxl.load_workbook(filename=file)
    except:
        logger.info("ERROR: Failed to access " + str(file))
        return -1
    try:
        ws = wb["from_geodata"]
    except:
        logger.info("ERROR: Sheet \'from_geodata\' is missing in " + str(file))
        return -1

    i_row = start_row
    for dct_key in data_dict.keys():
        ws[key_col + str(i_row)].value = dct_key
        ws[val_col + str(i_row)].value = data_dict[dct_key]
        i_row += 1

    wb.save(str(file))
    wb.close()
    logger.info(" -- OK (Write to workbook)\n")
