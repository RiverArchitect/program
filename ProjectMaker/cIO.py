#!/usr/bin/python
# Filename: IO.py
try:
    import os, sys, logging
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging).")
try:
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\\.site_packages\\openpyxl\\")
    import openpyxl as oxl  # modified package
except:
    print("ERROR: Could not load OPENPYXL. Modify module import references in cIO.py (sys.append path in line 8).")


class Read:
    def __init__(self, full_file_name):
        # type defines lines to read in .inp file
        self.logger = logging.getLogger("logfile_40")

        self.xlsx_file = full_file_name
        self.open_wb()

    def close_wb(self):
        try:
            self.wb.close()
        except:
            pass

    def open_wb(self):
        try:
            self.wb = oxl.load_workbook(filename=self.xlsx_file, read_only=True, data_only=True)
        except:
            self.wb = ""
            self.logger.info("ERROR: Failed to access " + str(self.xlsx_file).split("\\")[-1] + ".")
            self.logger.info("(full name: " + str(self.xlsx_file) + ")")
        try:
            self.ws = self.wb.worksheets[0]
        except:
            self.ws = []
            self.logger.info("ERROR: Could not find any worksheet.")

    def read_column(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        self.logger.info(
            "   * reading data column from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(column) + str(
                start_row) + ") ...")

        data = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                valid_content = False
            if valid_content:
                data.append(cell_value)
                __row__ += 1
        self.logger.info("   * ok")
        return data

    def read_row_str(self, row, start_col, **kwargs):
        # reads ROW beginning at COL until it meets an empty cell
        # row = INT
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
                if "end_col" in opt_var[0]:
                    # last relevant column
                    end_col = opt_var[1]
                if "if_row" in opt_var[0]:
                    # row that tells if the content of the actual cell is relevant
                    if_row = opt_var[1]
        except:
            pass

        if not ("col_skip" in locals()):
            col_skip = 1
        if not ("end_col" in locals()):
            end_col = "XFA"
        if not ("if_row" in locals()):
            if_row = row

        self.logger.info(
            "   * reading string row from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(start_col) + str(row) + ") ...")

        str_data = []
        valid_content = True
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        while valid_content:
            cell_string = str(self.ws[str(__col__) + str(row)].value)
            valid_content = self.test_cell_content(str(__col__) + str(if_row))
            if valid_content:
                str_data.append(cell_string)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
            else:
                valid_content = False
        self.logger.info("   * ok")
        return str_data

    def read_cell(self, col, row):
        cell_content = self.ws[str(col) + str(row)].value
        try:
            return float(cell_content)
        except:
            return cell_content

    def test_cell_content(self, cell_str):
        # cell_str = str, e.g., "C4"
        if str(self.ws[cell_str].value) == "None":
            return False
        else:
            return True

    @staticmethod
    def col_name_to_num(letters):
        pow = 1
        col_int = 0
        for letter in letters[::-1]:
            col_int += (int(letter, 36) - 9) * pow
            pow *= 26
        return col_int + 64

    @staticmethod
    def col_num_to_name(col_int):
        col_int -= 64
        letters = ''
        while col_int:
            mod = (col_int - 1) % 26
            letters += chr(mod + 65)
            col_int = (col_int - 1) // 26
        return ''.join(reversed(letters))

    def __call__(self):
        print("Class Info: <type> = Read (ProjectProposal)")


class Write:
    def __init__(self, template):
        self.logger = logging.getLogger("logfile_40")
        try:
            self.open_wb(str(template), 0)
        except:
            self.logger.info("ERROR: Could not open workbook (cIO.Write.__init__() -- " + str(template) + ").")

        self.wb_out_name = ""

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def open_wb(self, full_wb_path, ws):
        # full_wb_path = STR -- e.g., C:\\...\\my.xlsx
        # ws = STR or INT -- worksheet name or number in workbook
        self.logger.info("   * creating a copy of " + str(full_wb_path).split("\\")[-1])
        __wb__ = oxl.load_workbook(filename=full_wb_path)  # open template
        self.wb = __wb__  # copy template variable
        __wb__.close()  # close xlsx template

        # assign worksheet
        try:
            ws = int(ws)
        except:
            pass
        if type(ws) == int:
            try:
                self.ws = self.wb[str(self.wb.sheetnames[ws])]
            except:
                pass
        else:
            self.ws = self.wb[ws]

    def read_column(self, column, start_read_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = CHR, e.g., col = "B"
        # start_row = INT
        # returns dictionary with cell data and cell information

        data = {}
        valid_content = True
        __row__ = start_read_row
        while valid_content:
            try:
                cell_value = float(self.ws[str(column) + str(__row__)].value)
            except:
                valid_content = False
            if valid_content:
                data.update({cell_value: str(column) + str(__row__)})
                __row__ += 1
        self.logger.info("   * ok")
        return data

    def save_close_wb(self, *full_save_path):
        try:
            self.wb_out_name = full_save_path[0]
        except:
            pass

        try:
            self.logger.info("   * saving xlsx as:" + self.wb_out_name)
            self.wb.save(self.wb_out_name)
            self.wb.close()
            self.logger.info("   * ok")
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def write_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * WARNING: Could not write value to CELL " + str(column) + str(row))

    def write_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list must be a LIST object
        self.logger.info("   * writing column data (starting at " + str(column) + str(start_row) + ") ...")
        __row__ = start_row
        for val in data_list:
            self.ws[str(column) + str(__row__)].value = val
            __row__ += 1
        self.logger.info("   * ok")

    def write_row(self, start_col, row, data_list, **kwargs):
        # write a data list to ROW beginning at COL
        # data_list = LIST
        # row = INT
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
        except:
            pass

        if not ("col_skip" in locals()):
            col_skip = 1

        valid_content = True
        __col__ = start_col

        for e in data_list:
            try:
                self.write_cell(__col__, row, float(e))
            except:
                self.logger.info("   * WARNING: Non-numeric data in cell " + str(__col__) + str(row))
                self.write_cell(__col__, row, e)

            # update column to read next relevant row entry
            __col_num__ = Read.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip
            __col__ = Read.col_num_to_name(__col_num__)  # re-convert int to ascii-chr

    def __call__(self, *args, **kwargs):
        print("Class Info: <type> = Write (ProjectProposal)")

